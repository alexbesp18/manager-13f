import openpyxl

from manager13f import build, verify


def _sample_data():
    return {
        "meta": {
            "manager": "Example Capital LP",
            "cik": "0000000001",
            "latest_period": "2026-03-31",
            "latest_filed": "2026-05-15",
            "latest_accession": "0000000001-26-000001",
            "prior_period": "2025-12-31",
            "benchmark": "SPY",
            "benchmark_ytd": 6.5,
            "is_options_book": False,
            "n_positions": 2,
            "n_positions_prior": 1,
            "sleeves": {
                "long": 150_000_000,
                "call": 0,
                "put": 0,
                "gross": 150_000_000,
                "net_long": 150_000_000,
            },
            "sleeves_prior": {"long": 100_000_000, "call": 0, "put": 0, "gross": 100_000_000},
            "pareto_long": {
                "n": 2,
                "n80": 2,
                "n50": 1,
                "top5": 100.0,
                "top10": 100.0,
                "hhi": 5555.56,
                "effective_n": 1.8,
            },
            "unresolved_cusips": [],
            "enrich_errors": [],
            "generated": "2026-06-19T12:00:00+00:00",
        },
        "holdings": [
            {
                "side": "LONG",
                "ticker": "AAA",
                "name": "Alpha Automation",
                "sector": "Industrials",
                "shares": 1_000_000,
                "value": 100_000_000,
                "book_pct": 66.6667,
                "action": "New",
                "qoq_shares": None,
                "filing_px": 100.0,
                "price": 112.0,
                "post_filing_pct": 12.0,
                "ytd": 18.0,
                "rsi14": 62.0,
                "pct_vs_ma200": 9.0,
                "pct_from_hi": -5.0,
                "rating": "Buy",
                "n_analysts": 8,
                "target_mean": 125.0,
                "upside": 11.6,
            },
            {
                "side": "LONG",
                "ticker": "BBB",
                "name": "Beta Materials",
                "sector": "Materials",
                "shares": 500_000,
                "value": 50_000_000,
                "book_pct": 33.3333,
                "action": "Add",
                "qoq_shares": 25.0,
                "filing_px": 100.0,
                "price": 95.0,
                "post_filing_pct": -5.0,
                "ytd": -2.0,
                "rsi14": 41.0,
                "pct_vs_ma200": -4.0,
                "pct_from_hi": -20.0,
                "rating": "Hold",
                "n_analysts": 5,
                "target_mean": 100.0,
                "upside": 5.3,
            },
        ],
        "exits": [],
        "sectors": [],
    }


def test_build_preserves_13f_sheet_shape_and_structural_cleanliness(tmp_path):
    out = tmp_path / "manager.xlsx"

    build.build(_sample_data(), str(out))

    ok, detail = verify.structural_check(str(out))
    assert ok, detail

    wb = openpyxl.load_workbook(str(out), data_only=True)
    assert wb.sheetnames == ["13F"]
    ws = wb["13F"]
    assert ws.max_column == 22
    assert ws.cell(1, 1).value.startswith("EXAMPLE CAPITAL LP")
    assert ws.cell(1, 1).value.endswith("13F INTELLIGENCE")

    header_row = next(
        row for row in ws.iter_rows(values_only=True) if row and row[0] == "#" and row[1] == "Side"
    )
    assert header_row[:8] == ("#", "Side", "Ticker", "Company", "Sector", "Shares", "Value $", "% Book")
    assert "PostFil%" in header_row
    assert "Up%" in header_row
